import pprint

from openpyxl import load_workbook,workbook, worksheet
wb = load_workbook('emeadevit.xlsx')
#print(wb.sheetnames)


question_doc = {
    "Total" : 1516,
    "row" : 12,
    "column": 3,
    "Gender": {
        "Male" : 1015,
        "Female": 501,
    },
    "Age" : {
        "row": 12,
        "column":4,
        "16-24" : 84,
        "25-34" : 606,
        "35-44" : 610,
        "45-54" : 144,
        "55+"   : 72
    },
    "Country" : {
        "row": 12,
        "Column": 11,
        "Germany" : 513,
        "UK" : 502,
        "France" : 501,
    },
    "Industry Sector" : {
        "row": 12,
        "column": 14,
        "Architecture, Engineering & Building" : 72,
        "Arts & Culture" : 10,
        "Education" : 15,
        "Finance" : 77,
        "Healthcare": 38,
        "HR" : 13,
        "IT & Telecoms" : 1021,
        "Legal" : 5,
        "Manufacturing and Utilities" : 113,
        "Retail, Catering and Leisure": 71,
        "Sales, Media & Marketing" : 17,
        "Travel & Transport": 27,
        "Other": 42
    },
    "Company Size" : {
        "Sole Trader" : 8,
        "1-9 employees": 30,
        "10-49 employees": 61,
        "50-99 employees": 116,
        "100-249 employees": 116,
        "250-500 employees": 382,
        "More than 500 employees": 666,
    },
    "IT Decision Maker vs Developers" : {
        "IT Decision Maker" :756,
        "Developers": 760

    }
}

# for name in wb.sheetnames:
#     ws = wb[name]
#     question = ws["A10"].value
#     print(f"{question}")


def get_responses(ws:worksheet, row=16):

    responses = {}
    while True:
        response = ws.cell(row=row, column=2).value
        if response is None:
            break
        else:
            responses[response] = None
        row = row + 2
    return responses

def get_questions(ws:worksheet, row=16):

    question = {}

    while True:
        main_question = ws.cell( row=row, column=1).value
        if main_question is None:
            break
        else:
            question["question"] = {main_question:ws.cell(row=10, column=1).value}
            responses = get_responses(ws, row=row)
            question["responses"]=responses
            pprint.pprint(question)
            # k=input("Next..")
        row = row + len(responses) * 2 + 1

for name in wb.sheetnames:
    ws = wb[name]
    #print( ws.cell(row=16, column=2).value)
    print(f"Questions in sheet '{ws.title}'")
    get_questions(ws)
