from enum import Enum
import pprint

from openpyxl import load_workbook,workbook, worksheet


MainGroups={
 #   "Total" :3,
    "Gender": (4,2),
    "Age": (6,5),
    "Country": (11,3),
    "Industry Sector": (14, 13),
    "Company Size": (27,7),
    "IT Decision Maker vs Developers": (34,2)
}

SubGroups = {
    "Male" : 4,
    "Female" : 5,
    "16-24" : 6,
    "25-34" : 7,
    "35-44" : 8,
    "45-54" : 9,
    "55+" : 10,
    "Germany" : 11,
    "UK" : 12,
    "France" : 13,
    "Architecture Engineering & Building" : 14,
    "Arts & Culture" : 15,
    "Education" : 16,
    "Finance" : 17,
    "Healthcare" : 18,
    "HR" : 19,
    "IT & Telecoms": 20,
    "Legal" : 21,
    "Manufacturing & Utilities" : 22,
    "Retail, Catering & Leisure" : 23,
    "Sales, Media & Marketing" : 24,
    "Travel & Transport" : 25,
    "Other" : 26,
    "Sole Trader" : 27,
    "1 - 9 employees" : 28,
    "10 - 49 employees" : 29,
    "50 - 99 employees" : 30,
    "100 - 249 employees" : 31,
    "250 - 500 employees" : 32,
    "More than 500 employees" : 33,
    "IT Decision Maker" : 34,
    "Developers" : 35,
}



class CensusWorkbook:

    def __init__(self, excel_filename):
        self._workbook = load_workbook(filename=excel_filename, read_only=True)
        self._sheets = {}
        for name in self._workbook.sheetnames:
            self._sheets[name] = self._workbook[name]

    @property
    def workbook(self):
        return self._workbook

    @property
    def sheet_names(self):
        return list(self._sheets.keys())

    @property
    def sheets(self):
        return list(self._sheets.itervalues())

    def sheet(self, name):
        return self._sheets[name]

class CensusSheet:

    # def get_column_values(self, sheet,value_dict, column=2):
    #     response_values = {}
    #     for k, index in enumerate(value_dict.keys(), 16):
    #         value_dict[k] = sheet.cell(row=index, column=column).value
    #     for i in range(16, 16 + self.question_count() * 2, 2):
    #         response_values[(sheet.cell(row=i, column=column).value).values
    #     return response_values

    def __init__(self, workbook, name="None"):
        self._question_id = name
        self._workbook = workbook
        if self._question_id.startswith("Q"):
            self._sheet = self._workbook[self._question_id]
            self._responses = CensusSheet.parse_responses(self._sheet)
        else:
            raise ValueError(f"Bad name for sheet {name}, must begin with 'Q'")

    def sheet_name(self):
        return self._sheet.name

    @property
    def responses(self):
        return self._responses

    @property
    def question_id(self):
        return self._question_id

    @property
    def question(self):
        return self._sheet.cell(row=10, column=1).value

    @property
    def statement(self):
        return self._sheet.cell(row=11, column=1).value

    @staticmethod
    def parse_responses(sheet):
        """
        Read questions from row 16 column three
        Stop when we reach two blank lines which indicates
        the end of the questions
        :return: question list
        """
        row = 16
        column = 2
        responses = []
        while True:
            q = sheet.cell(row, column).value
            if q is None:
                break
            else:
                responses.append(q)
            row = row + 2

        return responses


    def response_offset(self, count=1):
            return 16 + ( 2 * count - 1)

    def column_value(self, question_number, column=1):
        row_offset = 16 + ( 2 * (question_number - 1))
        return self._sheet.cell(row=row_offset,column=column).value

    def column_values(self, column=1):
        for i in range(1, self.question_count + 1):
            yield self.column_value(i, column)

    def response_doc(self, main_group, sub_group, column_index):
        sub_group_dict = { sub_group :dict(list(zip(self.responses, self.column_values(column_index))))}
        return { main_group : sub_group_dict}

    def response_docs(self):
        top={self.question_id : self.question,
             "statement" : self.statement}

        for top_group, position in MainGroups.items():
            field_count = 0
            for field_name,column in SubGroups.items():
                rdoc= self.response_doc(top_group, field_name, column)
                field_count = field_count + 1
                if field_count > position[1]:
                    break

    @property
    def question_count(self):
        return len(self._responses)







if __name__ == "__main__":

    cb = CensusWorkbook("emeadevit.xlsx")
    print(cb.sheet_names)
    q3 = CensusSheet(cb.workbook, "Q3")
    print(q3.question)
    print(q3.responses)

    q =CensusSheet(cb.workbook, f"Q1")
    for doc in q.response_docs():
        pprint.pprint(doc)
