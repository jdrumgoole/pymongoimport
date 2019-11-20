import argparse
import os
import sys

import pymongo
from openpyxl import load_workbook,workbook, worksheet

class ExcelWorkbook:

    def __init__(self, excel_filename):
        self._workbook = load_workbook(filename=excel_filename, read_only=True)
        self._sheets = {}
        for name in self._workbook.sheetnames:
            self._sheets[name] = self._workbook[name]

    @property
    def workbook(self):
        return self._workbook

    @property
    def sheet_names(self):
        return list(self._sheets.keys())

    @property
    def sheets(self):
        return list(self._sheets.itervalues())

    def sheet(self, name):
        return self._sheets[name]


class ExcelSheet:

    def __init__(self, workbook, sheet_name="None"):

        self._workbook = workbook
        self._sheet_name = sheet_name
        if sheet_name:
            self._sheet = self._workbook[self._sheet_name]
        else:
            self._sheet = self._workbook.active

    def sheet_name(self):
        return self._sheet.name

    @property
    def sheet(self):
        return self._sheet

if __name__ == "__main__":

    parser = argparse.ArgumentParser()

    parser.add_argument("--host", default="mongodb://localhost:27017")
    parser.add_argument("--excelfile")
    parser.add_argument("--sheetname")
    parser.add_argument("--minrow", type=int)
    parser.add_argument("--mincol", type=int)
    parser.add_argument("--maxrow", type=int)
    parser.add_argument("--maxcol", type=int)
    parser.add_argument("--database", default="census")
    parser.add_argument("--collection", default="survey")
    parser.add_argument("--lowerright")
    parser.add_argument("--drop", action="store_true", default=False)
    args = parser.parse_args()

    client = pymongo.MongoClient(args.host)
    db = client[args.database]
    collection = db[args.collection]

    if args.drop:
        print(f"Dropping collection {args.collection}")
        db.drop_collection(args.collection)

    print(f"Writing data to cluster {args.host} collection:{args.database}.{args.collection}")

    if args.excelfile:
        if os.path.isfile(args.excelfile):
            cb = ExcelWorkbook(args.excelfile)
        else:
            print(f"{args.excelfile} is not a file")
            sys.exit(1)
    else:
        print(f'--excelfile is not specified')
        sys.exit(1)

    if args.sheetname:
        sh = ExcelSheet(cb.workbook, args.sheetname)
    else:
        print(f'--sheetname is not specified')
        sys.exit(1)



    doc = {}

    count = 1

    col_index = {}
    count = 1
    for row in sh.sheet.iter_rows(min_row=args.minrow,
                                   min_col=args.mincol,
                                   max_row=args.minrow,
                                   max_col=args.maxcol,
                                   values_only=True):

        for v in row:
            if v is None:
                continue
            else:

                col_index[count] = v
                print(f"{count}. {v}")
                count = count + 1


    row_index = {}
    count = 1
    for row in sh.sheet.iter_rows(min_row=args.minrow,
                                  min_col=args.mincol,
                                  max_row=args.maxrow,
                                  max_col=args.mincol,
                                  values_only=True):
        for v in row:
            if v is None:
                continue
            else:
                row_index[count] = v
                print(f"{count}. {v}")
                count = count + 1

    row_count = 1
    col_count = 1
    for row in sh.sheet.iter_rows(min_row=args.minrow+1,
                                  min_col=args.mincol+1,
                                  max_row=args.maxrow,
                                  max_col=args.maxcol,
                                  values_only=True):

        col_count=1
        for v in row:
            doc = {}
            result={}
            #doc[col_index[col_count]]=v
            #result[row_index[row_count]] = doc
            doc["title"] = row_index[row_count]
            doc[col_index[col_count]] = v
            print(doc)
            collection.insert_one(doc)
            col_count = col_count + 1

        row_count = row_count + 1

        # if count == 1 :
        #     for x_axis,i in enumerate(row, 1):
        #         if i == 1 :
        #             continue
        #         else:
        #             doc[x_axis]= {}
        #
        # else:
        #     for y_axis,i in enumerate(row, 1):
        #         if i == 1:
        #             doc[x_axis][y_axis ] = {}
        #         else:
        #             doc[x_axis][y_axis]






    # for row_count, row in enumerate(range, 1):
    #     for column_count, column in enumerate(row, 1):
    #         if row_count == 1 and column_count == 1 :
    #             continue
    #         elif row_count == 1:
    #             doc[column.value] = {}
    #         elif row_count > 1:
    #             doc
    #         doc[row.value][column.value] = {}
    #         if row_count > 1 and column_count > 1 :
    #             print(f" {row_count},{column_count} {row.value} ", end="")
    #
    #
    #
    #     print("")

    # db = client["census"]
    # collection = db["survey"]
    # for sheet_name in cb.sheet_names:
    #     cs = CensusSheet(cb.workbook, sheet_name)
    #     print(f"Processing sheet: {sheet_name}")
    #     collection.insert_one(cs.response_docs())

